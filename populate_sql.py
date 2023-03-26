import os
import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from movies.models import Movies
from admins.models import Admins


def populate():
    movie_data = read_xlsx()
    add_movie(movie_data)
    add_default_admin()


def add_default_admin():
    admin = [
        Admins(username='admin', password='123456')]
    # m = Admins.objects.get_or_create()[0]
    # m.save()
    Admins.objects.bulk_create(admin)


def add_movie(data):
    movie_list = [
        Movies(id=item['id'], title=item['title'], director=item['director'], date=item['date'], actor=item['actor'],
               score=item['score'], intro=item['intro'], tags=item['tags'],
               time_length=item['time_length'], pic=item['pic']) for item in data]
    # m = Movies.objects.get_or_create()[0]
    # m.save()
    Movies.objects.bulk_create(movie_list)
    return movie_list


def read_xlsx():
    # Load xlsx
    workbook = load_workbook(filename='movies.xlsx')

    # Select first sheet
    sheet = workbook.worksheets[0]

    # Get sheet headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)

    # Save data
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for i, value in enumerate(row):
            item[headers[i]] = value
        data.append(item)
    print(data)
    return data


if __name__ == '__main__':
    print('Starting Rango population script...')
    populate()
